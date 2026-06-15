"""
COG: Admin Log — /adminlog
Xuất log giao dịch phường thị + tặng quà ra file Excel.
Chỉ OWNER_ID mới dùng được.
"""
from __future__ import annotations
from typing import Any
import discord
from discord import app_commands
from discord.ext import commands
from utils.embeds import e_loi, e_ok, owner_only_check
from utils.config import OWNER_IDS
from utils.database import get_giao_dich_log
import io, time, datetime
from utils.embeds import safe_followup

# ══════════════════════════════════════════════════════
#  HELPER
# ══════════════════════════════════════════════════════

LOAI_LABEL = {
    "phien_cho": "Phường Thị",
    "tang_lt":   "Tặng LT",
    "tang_dan":  "Tặng Đan",
}

def _ts_to_str(ts: int) -> str:
    if not ts: return ""
    return datetime.datetime.fromtimestamp(ts, tz=datetime.timezone(datetime.timedelta(hours=7)))\
        .strftime("%Y-%m-%d %H:%M:%S")

async def _fetch_display_names(bot: discord.Client, user_ids: set[int]) -> dict[int, str]:
    """Lấy display name của từng user_id. Cache theo lần gọi."""
    result: dict[int, str] = {}
    for uid in user_ids:
        try:
            user = bot.get_user(uid) or await bot.fetch_user(uid)
            result[uid] = f"{user.display_name} ({user.name})" if user else f"[ID:{uid}]"
        except Exception:
            result[uid] = f"[ID:{uid}]"
    return result

async def _build_excel(rows: list[dict[str, Any]], bot: discord.Client) -> bytes:
    """Tạo file Excel từ list rows, trả về bytes."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        from openpyxl.worksheet.worksheet import Worksheet
    except ImportError:
        raise RuntimeError("Thiếu thư viện openpyxl! Chạy: pip install openpyxl")

    # Thu thập tất cả user_id cần resolve
    uid_set: set[int] = set()
    for r in rows:
        if r.get("sender_id"):   uid_set.add(r["sender_id"])
        if r.get("receiver_id"): uid_set.add(r["receiver_id"])
    names = await _fetch_display_names(bot, uid_set)

    wb = openpyxl.Workbook()

    # ── Sheet tổng hợp ──────────────────────────────────────────
    ws_all = wb.active
    ws_all.title = "Tất Cả"

    # ── Sheet theo loại ─────────────────────────────────────────
    sheets = {
        "phien_cho": wb.create_sheet("Phường Thị"),
        "tang_lt":   wb.create_sheet("Tặng LT"),
        "tang_dan":  wb.create_sheet("Tặng Đan"),
    }

    headers = [
        "ID", "Loại", "Thời Gian (GMT+7)",
        "Sender ID", "Sender",
        "Receiver ID", "Receiver",
        "Vật Phẩm", "Số Lượng", "LT Giao Dịch", "Ghi Chú"
    ]
    col_widths = [8, 14, 22, 18, 28, 18, 28, 24, 12, 16, 20]

    # Style header
    header_font  = Font(bold=True, color="FFFFFF", size=11)
    header_fill  = PatternFill("solid", fgColor="1F4E79")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin = Side(style="thin", color="AAAAAA")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    alt_fill = PatternFill("solid", fgColor="EEF4FB")

    def _setup_sheet(ws: Worksheet) -> None:
        ws.append(headers)
        for ci, (hdr, w) in enumerate(zip(headers, col_widths), 1):
            cell = ws.cell(row=1, column=ci)
            cell.font   = header_font
            cell.fill   = header_fill
            cell.alignment = header_align
            cell.border = border
            ws.column_dimensions[get_column_letter(ci)].width = w
        ws.row_dimensions[1].height = 30
        ws.freeze_panes = "A2"

    _setup_sheet(ws_all)
    for ws in sheets.values():
        _setup_sheet(ws)

    loai_fill = {
        "phien_cho": PatternFill("solid", fgColor="D6EAF8"),
        "tang_lt":   PatternFill("solid", fgColor="D5F5E3"),
        "tang_dan":  PatternFill("solid", fgColor="FEF9E7"),
    }

    def _write_row(ws: Worksheet, row_data: list[Any], row_idx: int, loai: str) -> None:
        for ci, val in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=ci, value=val)
            cell.border = border
            cell.alignment = Alignment(vertical="center", wrap_text=(ci in (5, 7, 8)))
            if row_idx % 2 == 0:
                cell.fill = loai_fill.get(loai, alt_fill)

    row_idx_all = 2
    row_idx_sheet = {k: 2 for k in sheets}

    # Tổng kê LT theo loại
    tong_lt = {"phien_cho": 0, "tang_lt": 0, "tang_dan": 0}

    for r in rows:
        loai = r.get("loai", "")
        sid  = r.get("sender_id", 0)
        rid  = r.get("receiver_id", 0)
        ts   = r.get("thoi_gian", 0)
        lt   = r.get("gia_lt", 0) or 0
        tong_lt[loai] = tong_lt.get(loai, 0) + lt

        row_data: list[Any] = [
            r.get("id", ""),
            LOAI_LABEL.get(loai, loai),
            _ts_to_str(ts),
            sid, names.get(sid, str(sid)),
            rid, names.get(rid, str(rid)),
            r.get("item_name", ""),
            r.get("so_luong", 1),
            lt,
            r.get("ghi_chu", ""),
        ]

        _write_row(ws_all, row_data, row_idx_all, loai)
        row_idx_all += 1

        if loai in sheets:
            _write_row(sheets[loai], row_data, row_idx_sheet[loai], loai)
            row_idx_sheet[loai] += 1

    # ── Sheet tổng kê ────────────────────────────────────────────
    ws_sum = wb.create_sheet("Tổng Kê", 0)  # đặt đầu tiên
    ws_sum.title = "Tổng Kê"
    ws_sum.column_dimensions["A"].width = 22
    ws_sum.column_dimensions["B"].width = 16
    ws_sum.column_dimensions["C"].width = 16

    sum_headers = ["Loại Giao Dịch", "Số Giao Dịch", "Tổng LT"]
    for ci, h in enumerate(sum_headers, 1):
        c = ws_sum.cell(row=1, column=ci, value=h)
        c.font = header_font; c.fill = header_fill
        c.alignment = header_align; c.border = border
    ws_sum.row_dimensions[1].height = 28

    counts = {loai: sum(1 for r in rows if r.get("loai") == loai)
              for loai in ["phien_cho", "tang_lt", "tang_dan"]}

    for ri, loai in enumerate(["phien_cho", "tang_lt", "tang_dan"], 2):
        row: list[Any] = [LOAI_LABEL.get(loai, loai), counts[loai], tong_lt.get(loai, 0)]
        fill = loai_fill.get(loai, alt_fill)
        for ci, val in enumerate(row, 1):
            c = ws_sum.cell(row=ri, column=ci, value=val)
            c.fill = fill; c.border = border
            c.alignment = Alignment(horizontal="center", vertical="center")

    # Tổng cộng
    total_row: list[Any] = ["TỔNG CỘNG", sum(counts.values()), sum(tong_lt.values())]
    total_fill = PatternFill("solid", fgColor="1F4E79")
    for ci, val in enumerate(total_row, 1):
        c = ws_sum.cell(row=5, column=ci, value=val)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = total_fill; c.border = border
        c.alignment = Alignment(horizontal="center", vertical="center")

    # Metadata
    ws_sum.cell(row=7, column=1, value="Xuất lúc:").font = Font(italic=True)
    ws_sum.cell(row=7, column=2, value=_ts_to_str(int(time.time())))
    ws_sum.cell(row=8, column=1, value="Tổng bản ghi:").font = Font(italic=True)
    ws_sum.cell(row=8, column=2, value=len(rows))

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


# ══════════════════════════════════════════════════════
#  COG
# ══════════════════════════════════════════════════════

class AdminLogCog(commands.Cog):
    def __init__(self, bot: commands.Bot):
        self.bot = bot

    @app_commands.command(
        name="adminlog",
        description="[Admin] Xuất log giao dịch phường thị & tặng quà ra Excel")
    @app_commands.describe(
        loai="Lọc theo loại (bỏ trống = tất cả)",
        user_id="Lọc theo Discord user ID (bỏ trống = tất cả)",
        limit="Số bản ghi tối đa (mặc định 2000)")
    @app_commands.choices(loai=[
        app_commands.Choice(name="Tất cả",      value="all"),
        app_commands.Choice(name="Phường Thị",  value="phien_cho"),
        app_commands.Choice(name="Tặng LT",     value="tang_lt"),
        app_commands.Choice(name="Tặng Đan",    value="tang_dan"),
    ])
    @owner_only_check(OWNER_IDS)
    async def adminlog(
        self,
        inter: discord.Interaction,
        loai: app_commands.Choice[str] | None = None,
        user_id: str | None = None,
        limit: int = 2000,
    ) -> None:
        await inter.response.defer(ephemeral=True, thinking=True)

        # Parse user_id
        uid: int | None = None
        if user_id:
            try:
                uid = int(user_id.strip())
            except ValueError:
                return await safe_followup(inter, 
                    embed=e_loi("❌ User ID không hợp lệ", "Vui lòng nhập số Discord ID."),
                    ephemeral=True)

        loai_val: str | None = None if (loai is None or loai.value == "all") else loai.value
        limit = max(1, min(limit, 5000))

        rows: list[dict[str, Any]] = await get_giao_dich_log(user_id=uid, loai=loai_val, limit=limit)

        if not rows:
            return await safe_followup(inter, 
                embed=e_ok("📋 Không có dữ liệu",
                    f"Không tìm thấy giao dịch nào" +
                    (f" cho user `{uid}`" if uid else "") +
                    (f" loại `{loai_val}`" if loai_val else "") + "."),
                ephemeral=True)

        try:
            excel_bytes = await _build_excel(rows, self.bot)
        except RuntimeError as e:
            return await safe_followup(inter, 
                embed=e_loi("❌ Lỗi xuất Excel", str(e)), ephemeral=True)

        now_str = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"giao_dich_log_{now_str}.xlsx"
        file = discord.File(io.BytesIO(excel_bytes), filename=filename)

        loai_str  = LOAI_LABEL.get(loai_val, "Tất Cả") if loai_val else "Tất Cả"
        uid_str   = f" | User: `{uid}`" if uid else ""
        count_str = f"**{len(rows)}** bản ghi"

        await safe_followup(inter, 
            embed=discord.Embed(
                title="📊 Xuất Log Giao Dịch",
                description=(
                    f"✅ Xuất thành công {count_str}\n"
                    f"📂 Loại: **{loai_str}**{uid_str}\n"
                    f"📅 Thời gian: `{_ts_to_str(int(time.time()))}`"
                ),
                color=0x1F4E79),
            file=file,
            ephemeral=True)


async def setup(bot: commands.Bot):
    await bot.add_cog(AdminLogCog(bot))
